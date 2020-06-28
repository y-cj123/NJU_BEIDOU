# -*- coding: utf-8 -*-
"""
Created on Tue Mar  5 22:29:02 2019

@author: p7nkpaNda
"""

import xlrd

import difflib

from openpyxl import load_workbook


if __name__ == '__main__':
    wb = load_workbook(u'1.xlsx')
    ws = wb[u"实际数据源"]
    execlfile = xlrd.open_workbook(u'1.xlsx')
    sheet = execlfile.sheet_by_name(u"实际数据源")
    cols2 = sheet.col_values(3)
    a = sheet.col_values(0)

    date='test'

    for i in range(1,len(a)):
        if a[i]=='':
            break
        Max=['','',0]
        count=0
        for j in range(1,len(cols2)):
            rate = difflib.SequenceMatcher(None, a[i], cols2[j]).quick_ratio()
            if rate > Max[-1] and rate >=0.1:
                Max = [a[i],cols2[j],rate]
                count = j
        ws['B'+str(i+1)] = Max[1]
        ws["C"+str(i+1)] = ws.cell(row = count+1,column =5).value


    wb.save(u'1.xlsx')

